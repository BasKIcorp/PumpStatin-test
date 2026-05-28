"""Админ API: пользователи, профили/фронт, PDF, БД, плагины."""

import io
import re
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import delete, select, text
from sqlalchemy.exc import SQLAlchemyError

from app.api.deps import require_admin
from app.core.config import settings
from app.core.profile_loader import list_profiles
from app.db.models import CatalogItemModel, PumpModel
from app.db.session import SessionLocal
from app.services import config_store
from app.services.accounts import create_user, delete_user, list_users_admin, update_user

router = APIRouter()


# --- Users ---


class UserCreateBody(BaseModel):
    username: str
    password: str
    displayName: str
    profileId: str
    organization: str | None = None
    role: str = "user"


class UserUpdateBody(BaseModel):
    password: str | None = None
    displayName: str | None = None
    profileId: str | None = None
    organization: str | None = None
    role: str | None = None


@router.get("/meta")
def admin_meta(_: Annotated[dict, Depends(require_admin)]):
    return config_store.admin_meta()


@router.get("/users")
def admin_list_users(_: Annotated[dict, Depends(require_admin)]):
    return {"users": list_users_admin()}


@router.post("/users")
def admin_create_user(body: UserCreateBody, admin: Annotated[dict, Depends(require_admin)]):
    try:
        user = create_user(
            username=body.username,
            password=body.password,
            display_name=body.displayName,
            profile_id=body.profileId,
            organization=body.organization,
            role=body.role,
        )
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    return {"user": user}


@router.put("/users/{username}")
def admin_update_user(
    username: str,
    body: UserUpdateBody,
    admin: Annotated[dict, Depends(require_admin)],
):
    try:
        user = update_user(
            username,
            password=body.password,
            display_name=body.displayName,
            profile_id=body.profileId,
            organization=body.organization,
            role=body.role,
        )
    except ValueError as e:
        raise HTTPException(404, str(e)) from e
    return {"user": user}


@router.delete("/users/{username}")
def admin_delete_user(username: str, admin: Annotated[dict, Depends(require_admin)]):
    if username == admin["username"]:
        raise HTTPException(400, "Cannot delete yourself")
    try:
        delete_user(username)
    except ValueError as e:
        raise HTTPException(404, str(e)) from e
    return {"ok": True}


# --- Profiles / frontend / plugins ---


class ProfilePluginsBody(BaseModel):
    displayName: str | None = None
    theme: str | None = None
    algorithm: str | None = None
    database: str | None = None
    pdfTemplate: str | None = None
    layoutVariant: str | None = None
    active: bool | None = None


class BrandingBody(BaseModel):
    branding: dict[str, Any]


@router.get("/profiles")
def admin_profiles(_: Annotated[dict, Depends(require_admin)]):
    registry = config_store.load_registry().get("profiles", [])
    items = []
    for entry in registry:
        pid = entry["id"]
        try:
            profile = config_store.load_profile_yaml(pid)
        except FileNotFoundError:
            profile = {}
        items.append(
            {
                "registry": entry,
                "profile": {
                    "id": profile.get("id", pid),
                    "displayName": profile.get("displayName", entry.get("displayName")),
                    "theme": profile.get("theme"),
                    "algorithm": profile.get("algorithm"),
                    "database": profile.get("database"),
                    "pdfTemplate": profile.get("pdfTemplate"),
                    "active": profile.get("active", True),
                },
            }
        )
    return {"profiles": items}


@router.get("/profiles/{profile_id}")
def admin_profile_detail(profile_id: str, _: Annotated[dict, Depends(require_admin)]):
    try:
        profile = config_store.load_profile_yaml(profile_id)
        branding = config_store.load_branding_yaml(profile_id)
    except FileNotFoundError:
        raise HTTPException(404, "Profile not found") from None
    return {
        "registry": config_store.get_registry_entry(profile_id),
        "profile": profile,
        "branding": branding,
    }


@router.put("/profiles/{profile_id}/plugins")
def admin_update_profile_plugins(
    profile_id: str,
    body: ProfilePluginsBody,
    _: Annotated[dict, Depends(require_admin)],
):
    try:
        profile = config_store.load_profile_yaml(profile_id)
    except FileNotFoundError:
        raise HTTPException(404, "Profile not found") from None

    patch = body.model_dump(exclude_none=True)
    for key in ("displayName", "theme", "algorithm", "database", "pdfTemplate", "active"):
        if key in patch:
            profile[key] = patch[key]

    try:
        config_store.save_profile_yaml(profile_id, profile)
    except ValueError as e:
        raise HTTPException(400, str(e)) from e

    registry_patch: dict[str, Any] = {}
    if body.displayName is not None:
        registry_patch["displayName"] = body.displayName
    if body.theme is not None:
        registry_patch["theme"] = body.theme
    if body.pdfTemplate is not None:
        registry_patch["pdfTemplate"] = body.pdfTemplate
    if body.layoutVariant is not None:
        registry_patch["layoutVariant"] = body.layoutVariant
    if registry_patch:
        config_store.upsert_registry_entry(profile_id, registry_patch)

    if body.layoutVariant is not None:
        try:
            branding = config_store.load_branding_yaml(profile_id)
            branding["layoutVariant"] = body.layoutVariant
            config_store.save_branding_yaml(profile_id, branding)
        except FileNotFoundError:
            pass

    try:
        profile = config_store.load_profile_yaml(profile_id)
        branding = config_store.load_branding_yaml(profile_id)
    except FileNotFoundError:
        raise HTTPException(404, "Profile not found") from None
    return {
        "registry": config_store.get_registry_entry(profile_id),
        "profile": profile,
        "branding": branding,
    }


@router.put("/profiles/{profile_id}/branding")
def admin_update_branding(
    profile_id: str,
    body: BrandingBody,
    _: Annotated[dict, Depends(require_admin)],
):
    if profile_id not in config_store.list_profile_ids():
        raise HTTPException(404, "Profile not found")
    config_store.save_branding_yaml(profile_id, body.branding)
    return {"branding": body.branding}


# --- Database ---


def _require_postgres_admin():
    if settings.use_mock_db:
        raise HTTPException(
            503,
            "Редактирование БД недоступно при USE_MOCK_DB=true. "
            "Установите USE_MOCK_DB=false и перезапустите API.",
        )


class PumpBody(BaseModel):
    id: str
    product_line: str
    name: str
    nominal_flow: float
    nominal_head: float
    power_kw: float | None = None


class CatalogItemBody(BaseModel):
    source_key: str
    value: str
    label: str


class TableColumnBody(BaseModel):
    name: str
    type: str
    nullable: bool = True
    primary_key: bool = False


class CreateTableBody(BaseModel):
    table_name: str
    columns: list[TableColumnBody]


class AlterTableBody(BaseModel):
    table_name: str
    action: str
    column_name: str | None = None
    new_column_name: str | None = None
    column_type: str | None = None
    nullable: bool | None = None


_IDENT_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
_TYPE_ALLOWED = {
    "text": "TEXT",
    "varchar": "VARCHAR(255)",
    "integer": "INTEGER",
    "bigint": "BIGINT",
    "numeric": "NUMERIC",
    "double": "DOUBLE PRECISION",
    "boolean": "BOOLEAN",
    "timestamp": "TIMESTAMP",
    "jsonb": "JSONB",
}


def _safe_ident(value: str, field_name: str) -> str:
    if not value or not _IDENT_RE.match(value):
        raise HTTPException(400, f"Invalid {field_name}")
    return value


def _safe_type(value: str) -> str:
    key = (value or "").strip().lower()
    if key not in _TYPE_ALLOWED:
        raise HTTPException(400, f"Unsupported column type: {value}")
    return _TYPE_ALLOWED[key]


@router.get("/database/status")
async def admin_db_status(_: Annotated[dict, Depends(require_admin)]):
    if settings.use_mock_db:
        return {"mode": "mock", "editable": False}
    async with SessionLocal() as session:
        pumps = await session.execute(select(PumpModel.id))
        catalog = await session.execute(select(CatalogItemModel.id))
        return {
            "mode": "postgres",
            "editable": True,
            "pumpCount": len(pumps.scalars().all()),
            "catalogItemCount": len(catalog.scalars().all()),
        }


@router.get("/database/schema")
async def admin_db_schema(_: Annotated[dict, Depends(require_admin)]):
    _require_postgres_admin()
    async with SessionLocal() as session:
        table_rows = await session.execute(
            text(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                ORDER BY table_name
                """
            )
        )
        tables: list[dict[str, Any]] = []
        for (table_name,) in table_rows.all():
            columns_rows = await session.execute(
                text(
                    """
                    SELECT column_name, data_type, is_nullable
                    FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = :table_name
                    ORDER BY ordinal_position
                    """
                ),
                {"table_name": table_name},
            )
            pk_rows = await session.execute(
                text(
                    """
                    SELECT kcu.column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu
                      ON tc.constraint_name = kcu.constraint_name
                     AND tc.table_schema = kcu.table_schema
                    WHERE tc.table_schema='public'
                      AND tc.table_name=:table_name
                      AND tc.constraint_type='PRIMARY KEY'
                    """
                ),
                {"table_name": table_name},
            )
            pk_cols = {name for (name,) in pk_rows.all()}
            tables.append(
                {
                    "name": table_name,
                    "columns": [
                        {
                            "name": c_name,
                            "type": c_type,
                            "nullable": is_nullable == "YES",
                            "primary_key": c_name in pk_cols,
                        }
                        for (c_name, c_type, is_nullable) in columns_rows.all()
                    ],
                }
            )
    return {"tables": tables}


@router.post("/database/schema/create-table")
async def admin_db_create_table(
    body: CreateTableBody, _: Annotated[dict, Depends(require_admin)]
):
    _require_postgres_admin()
    table_name = _safe_ident(body.table_name, "table_name")
    if not body.columns:
        raise HTTPException(400, "At least one column is required")
    column_defs: list[str] = []
    pk_cols: list[str] = []
    for column in body.columns:
        name = _safe_ident(column.name, "column.name")
        sql_type = _safe_type(column.type)
        nullable_sql = "" if column.nullable else " NOT NULL"
        column_defs.append(f'"{name}" {sql_type}{nullable_sql}')
        if column.primary_key:
            pk_cols.append(name)
    if pk_cols:
        quoted_pk = ", ".join(f'"{c}"' for c in pk_cols)
        column_defs.append(f"PRIMARY KEY ({quoted_pk})")
    sql = f'CREATE TABLE "{table_name}" ({", ".join(column_defs)})'
    async with SessionLocal() as session:
        try:
            await session.execute(text(sql))
            await session.commit()
        except SQLAlchemyError as e:
            await session.rollback()
            raise HTTPException(400, f"Create table failed: {e}") from e
    return {"ok": True}


@router.post("/database/schema/alter-table")
async def admin_db_alter_table(
    body: AlterTableBody, _: Annotated[dict, Depends(require_admin)]
):
    _require_postgres_admin()
    table_name = _safe_ident(body.table_name, "table_name")
    action = (body.action or "").strip().lower()
    if action == "add_column":
        if not body.column_name or not body.column_type:
            raise HTTPException(400, "column_name and column_type are required")
        col_name = _safe_ident(body.column_name, "column_name")
        col_type = _safe_type(body.column_type)
        nullable_sql = "" if (body.nullable is not False) else " NOT NULL"
        sql = f'ALTER TABLE "{table_name}" ADD COLUMN "{col_name}" {col_type}{nullable_sql}'
    elif action == "rename_column":
        if not body.column_name or not body.new_column_name:
            raise HTTPException(400, "column_name and new_column_name are required")
        old_name = _safe_ident(body.column_name, "column_name")
        new_name = _safe_ident(body.new_column_name, "new_column_name")
        sql = f'ALTER TABLE "{table_name}" RENAME COLUMN "{old_name}" TO "{new_name}"'
    elif action == "drop_column":
        if not body.column_name:
            raise HTTPException(400, "column_name is required")
        col_name = _safe_ident(body.column_name, "column_name")
        sql = f'ALTER TABLE "{table_name}" DROP COLUMN "{col_name}"'
    else:
        raise HTTPException(400, "Unsupported action")
    async with SessionLocal() as session:
        try:
            await session.execute(text(sql))
            await session.commit()
        except SQLAlchemyError as e:
            await session.rollback()
            raise HTTPException(400, f"Alter table failed: {e}") from e
    return {"ok": True}


@router.post("/database/import/excel")
async def admin_db_import_excel(
    table_name: Annotated[str, Form(...)],
    file: Annotated[UploadFile, File(...)],
    _: Annotated[dict, Depends(require_admin)],
):
    _require_postgres_admin()
    safe_table = _safe_ident(table_name, "table_name")
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Only .xlsx/.xlsm files are supported")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid Excel file: {e}") from e
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "Excel is empty") from None
    headers = [str(h).strip() if h is not None else "" for h in header_raw]
    if not headers or any(not h for h in headers):
        raise HTTPException(400, "Header row must contain non-empty column names")
    headers = [_safe_ident(h, "header column") for h in headers]
    inserted = 0
    skipped = 0
    errors: list[str] = []
    col_sql = ", ".join(f'"{h}"' for h in headers)
    values_sql = ", ".join(f":{h}" for h in headers)
    sql = text(f'INSERT INTO "{safe_table}" ({col_sql}) VALUES ({values_sql})')
    async with SessionLocal() as session:
        for row_idx, row_values in enumerate(rows_iter, start=2):
            if row_values is None:
                continue
            row_list = list(row_values)
            if all(v is None or str(v).strip() == "" for v in row_list):
                skipped += 1
                continue
            try:
                params = {
                    headers[i]: (row_list[i] if i < len(row_list) else None)
                    for i in range(len(headers))
                }
                await session.execute(sql, params)
                inserted += 1
            except SQLAlchemyError as e:
                errors.append(f"Row {row_idx}: {e}")
        if inserted > 0:
            await session.commit()
        else:
            await session.rollback()
    return {"inserted": inserted, "skipped": skipped, "errors": errors[:50]}


@router.get("/database/pumps")
async def admin_list_pumps(_: Annotated[dict, Depends(require_admin)]):
    _require_postgres_admin()
    async with SessionLocal() as session:
        result = await session.execute(select(PumpModel))
        rows = result.scalars().all()
        return {
            "pumps": [
                {
                    "id": r.id,
                    "product_line": r.product_line,
                    "name": r.name,
                    "nominal_flow": r.nominal_flow,
                    "nominal_head": r.nominal_head,
                    "power_kw": r.power_kw,
                }
                for r in rows
            ]
        }


@router.post("/database/pumps")
async def admin_create_pump(body: PumpBody, _: Annotated[dict, Depends(require_admin)]):
    _require_postgres_admin()
    async with SessionLocal() as session:
        existing = await session.get(PumpModel, body.id)
        if existing:
            raise HTTPException(400, "Pump id already exists")
        session.add(
            PumpModel(
                id=body.id,
                product_line=body.product_line,
                name=body.name,
                nominal_flow=body.nominal_flow,
                nominal_head=body.nominal_head,
                power_kw=body.power_kw,
            )
        )
        await session.commit()
    return {"ok": True}


@router.put("/database/pumps/{pump_id}")
async def admin_update_pump(
    pump_id: str, body: PumpBody, _: Annotated[dict, Depends(require_admin)]
):
    _require_postgres_admin()
    async with SessionLocal() as session:
        row = await session.get(PumpModel, pump_id)
        if not row:
            raise HTTPException(404, "Pump not found")
        row.product_line = body.product_line
        row.name = body.name
        row.nominal_flow = body.nominal_flow
        row.nominal_head = body.nominal_head
        row.power_kw = body.power_kw
        await session.commit()
    return {"ok": True}


@router.delete("/database/pumps/{pump_id}")
async def admin_delete_pump(pump_id: str, _: Annotated[dict, Depends(require_admin)]):
    _require_postgres_admin()
    async with SessionLocal() as session:
        await session.execute(delete(PumpModel).where(PumpModel.id == pump_id))
        await session.commit()
    return {"ok": True}


@router.get("/database/catalog")
async def admin_list_catalog(
    source_key: str | None = None,
    _: Annotated[dict, Depends(require_admin)] = None,
):
    _require_postgres_admin()
    async with SessionLocal() as session:
        q = select(CatalogItemModel)
        if source_key:
            q = q.where(CatalogItemModel.source_key == source_key)
        result = await session.execute(q)
        return {
            "items": [
                {
                    "id": r.id,
                    "source_key": r.source_key,
                    "value": r.value,
                    "label": r.label,
                }
                for r in result.scalars().all()
            ]
        }


@router.post("/database/catalog")
async def admin_create_catalog_item(
    body: CatalogItemBody, _: Annotated[dict, Depends(require_admin)]
):
    _require_postgres_admin()
    async with SessionLocal() as session:
        session.add(
            CatalogItemModel(
                source_key=body.source_key,
                value=body.value,
                label=body.label,
            )
        )
        await session.commit()
    return {"ok": True}


@router.delete("/database/catalog/{item_id}")
async def admin_delete_catalog_item(
    item_id: int, _: Annotated[dict, Depends(require_admin)]
):
    _require_postgres_admin()
    async with SessionLocal() as session:
        await session.execute(
            delete(CatalogItemModel).where(CatalogItemModel.id == item_id)
        )
        await session.commit()
    return {"ok": True}


@router.get("/profiles-list")
def admin_profiles_simple(_: Annotated[dict, Depends(require_admin)]):
    """Справочник profileId для форм пользователей."""
    return {"profiles": list_profiles()}
